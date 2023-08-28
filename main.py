import openpyxl
import time
from multiprocessing import Pool


start_time = time.time()

wb = openpyxl.load_workbook(filename = 'C:/Users/cnxxd/PycharmProjects/task2/time intervals (practice).xlsx')
sheet = wb['Лист2']

sessions = {}
for i in range(2, 41523):
    sessions[i] = None

for i in range(2, 41523):
    sessions[i] = sheet[f'A{i}':f'D{i}']


list_values_date_dict = {}
for i in range(2, 41523):
    list_values_date_dict[i] = None

for i in range(2, 41523):
    list_values_date = []
    date = sessions[i]
    plus_date = date[0]
    normal_date = str(plus_date[1].value)
    list_values_date = normal_date[0:10].split('-')
    list_values_date_dict[i] = list_values_date


list_values_begin_dict = {}
for i in range(2, 41523):
    list_values_begin_dict[i] = None

for i in range(2, 41523):
    list_values_begin = []
    begin = sessions[i]
    plus_begin = begin[0]
    normal_begin = str(plus_begin[2].value)
    list_values_begin = normal_begin.split(':')
    time_begin = int(list_values_begin[0])*3600 + int(list_values_begin[1])*60 + int(list_values_begin[2])
    list_values_begin_dict[i] = time_begin

list_values_end_dict = {}
for i in range(2, 41523):
    list_values_end_dict[i] = None

for i in range(2, 41523):
    list_values_end = []
    end = sessions[i]
    plus_end = end[0]
    normal_end = str(plus_end[3].value)
    list_values_end = normal_end.split(':')
    time_end = int(list_values_end[0])*3600 + int(list_values_end[1])*60 + int(list_values_end[2])
    list_values_end_dict[i] = time_end

#counter_dict = {}
#for i in range(0, 86401):
    #counter_dict[i] = None

def go(a, b, list_values_date_dict, list_values_begin_dict, list_values_end_dict):
    counter_dict_1 = {}
    for i in range(0, 86401):
        counter_dict_1[i] = None
    for i in range(a, b):   # 21600|86400
        counter = 0
        for c in range(2, 41523):
            lol = list_values_date_dict[c]
            if lol[2] == '06':
                if ((int(list_values_begin_dict[c])<=i) and (int(list_values_end_dict[c])>=i)):
                    counter += 1
        counter_dict_1[i] = counter
    return counter_dict_1

def time_begin_end(wb, list_values_end_dict):
    sovp = {}
    for i in range(0, 86400):
        sovp[i] = None

    for i in range(21600, 86399):
        a = 0
        b = 0
        if counter_dict[i] is not None:
            a = counter_dict[i]
        if counter_dict[i + 1] is not None:
            b = counter_dict[i + 1]
        if (a or b) != 0:
            if a < b:
                sovp[i + 1] = b
            elif a > b:
                sovp[i] = a

    final_time_begin = {}
    for i in range(0, 86400):
        final_time_begin[i] = None
    counter = 0
    cnt_1 = {}
    for i in range(0, 86400):
        cnt_1[i] = None
    c = 0

    for i in range(21600, 86400):
        str_time = ''
        if (counter != sovp[i]) and (sovp[i] is not None):
            counter = sovp[i]
            str_time = f'{i // 3600}:{i // 60 % 60}:{i % 60}'
            final_time_begin[c] = str_time
            cnt_1[c] = counter
            c += 1

    final_time_end = {}
    for i in range(0, 86400):
        final_time_end[i] = None
    counter = 0
    cnt_2 = {}
    for i in range(0, 86400):
        cnt_2[i] = None
    c = 0

    for i in range(21600, 86400):
        str_time = ''
        if (counter != sovp[i]) and (sovp[i] is not None):
            counter = sovp[i]
            if i%60 == 0:
                str_time = f'{i // 3600}:{i // 60 % 60 - 1}:{59}'
            elif i%60 != 0:
                str_time = f'{i // 3600}:{i // 60 % 60}:{i % 60 - 1}'
            final_time_end[c] = str_time
            cnt_2[c] = counter
            c += 1

    f = open('test.txt', 'w')
    f.write(str(final_time_begin))
    f.close()

    f_1 = open('test2.txt', 'w')
    f_1.write(str(final_time_end))
    f.close()

    ws2 = wb.create_sheet(title='Itog')
    for row in range(2, 4000):
        for col in range(1, 4):
            if col == 1:
                _ = ws2.cell(column=col, row=row, value=final_time_begin[row - 2])
            if col == 2:
                _ = ws2.cell(column=col, row=row, value=final_time_end[row - 1])
            if col == 3:
                _ = ws2.cell(column=col, row=row, value=cnt_1[row - 2])
        _ = ws2.cell(column=2, row=3469, value=f'{list_values_end_dict[41522]//3600}:{list_values_end_dict[41522]//60%60}:{list_values_end_dict[41522]%60}')
    for row in range(2, 3470):
        _ = ws2.cell(column=4, row=row, value='06.06.2022')
    wb.save(filename = 'C:/Users/cnxxd/PycharmProjects/task2/time intervals (practice).xlsx')



if __name__ == '__main__':
    counter_dict = {}
    for i in range(0, 86401):
        counter_dict[i] = None

    with Pool(7) as pool:
        res = pool.starmap(go, [(21600, 30000, list_values_date_dict, list_values_begin_dict, list_values_end_dict), (30001, 40000, list_values_date_dict, list_values_begin_dict, list_values_end_dict), (40001, 50000, list_values_date_dict, list_values_begin_dict, list_values_end_dict), (50001, 60000, list_values_date_dict, list_values_begin_dict, list_values_end_dict), (60001, 70000, list_values_date_dict, list_values_begin_dict, list_values_end_dict), (70001, 80000, list_values_date_dict, list_values_begin_dict, list_values_end_dict), (80001, 86400, list_values_date_dict, list_values_begin_dict, list_values_end_dict),])

        counter_dict.update((k, v) for k, v in res[0].items() if v is not None)
        counter_dict.update((k, v) for k, v in res[1].items() if v is not None)
        counter_dict.update((k, v) for k, v in res[2].items() if v is not None)
        counter_dict.update((k, v) for k, v in res[3].items() if v is not None)
        counter_dict.update((k, v) for k, v in res[4].items() if v is not None)
        counter_dict.update((k, v) for k, v in res[5].items() if v is not None)
        counter_dict.update((k, v) for k, v in res[6].items() if v is not None)

        f = open('test.txt', 'w')
        f.write(str(counter_dict))
        f.close()

        #time_begin_end(wb, list_values_end_dict)



        print("--- %s seconds ---" % (time.time() - start_time))





